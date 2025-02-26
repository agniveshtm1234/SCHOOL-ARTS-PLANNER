while True:
    try:
        import mysql.connector as mc
        import os      #importing the required modules
        import cv2
        import time
        import subprocess as sb
        import tkinter as tk
        from copy import copy
        from tkinter import filedialog
        from tkinter import messagebox
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
        
        ab=input('Enter username of the mysql server::')
        bc=input('Enter password of the mysql server::')
        z=mc.connect(host='localhost',user=ab,passwd=bc)   #establishing connection with mysql
        c=z.cursor()
        data= 'SCHOOLARTS'  
        print()
        dl="show databases like '{}'".format(data)
        c.execute(dl)         
        y=c.fetchone()    #creating a database if it doesnt exist!
        if y is  None:
            print('Database Not Found! Creating a New Database.')
            dll="create database {}".format(data)
            c.execute(dll)
            z.database=data
        else:
            z.database=data

        schoolname=input('Enter The Name of Your School::')
        arname=input('Enter The Name of your Artsfest::')
        print()
        print("Upload The School Logo!!!")
        while True:
            root = tk.Tk()
            root.attributes('-topmost',True)
            root.withdraw()
            logo_path = filedialog.askopenfilename(title="SELECT SCHOOL LOGO") #showing the filedialogbox to upload school logo
            root.destroy()
            if logo_path == '':
                root = tk.Tk()
                root.attributes('-topmost',True)
                root.withdraw()                          #showing the messagebox to give error message
                a = messagebox.showerror('SCHOOL ARTS PLANNER','LOGO NOT RECIEVED\nTRY AGAIN!!!')
                root.destroy()
                continue
            elif not logo_path.endswith((".jpg",".png")):
                root = tk.Tk()
                root.attributes('-topmost',True)        
                root.withdraw()                          #showing the messagebox to give error message
                a = messagebox.showerror('SCHOOL ARTS PLANNER','UPLOAD IMAGES ONLY!!!\nTRY AGAIN!!!')
                root.destroy()
                continue
            else:
                print()
                print("Uploading the School Logo", end=" ", flush=True)
                for _ in range(10):
                    time.sleep(0.1)
                    print("██", end="", flush=True)
                print("100%")
                print("Uploaded Successfully!!!")
                break
        
        folder_path = os.path.abspath("..\\SCHOOL ARTS PLANNER\\..")
        print(folder_path)
        pcert_path = r"{}\\cert\\participation_certificates".format(folder_path)
        wcert_path = r"{}\\cert\\winners_certificates".format(folder_path)
        excel_path = r"{}\\INFO".format(folder_path)
        qu = [pcert_path,wcert_path,excel_path]
        for i in qu:
            if not os.path.exists(i):
                os.makedirs(pcert_path)        #making a folder if the folder doesnt exists
                os.makedirs(wcert_path)
                os.makedirs(excel_path)
          
        def register():
            a=int(input('Enter Item ID:'))
            b=input('Enter Item Name:')
            print()
            print('press 1 for offstage type:')
            print('press 2 for onstage type:')
            ch=int(input('Enter your choice ( 1 or 2):'))      
            if ch==1:
                d='offstage'
            elif ch==2:
                d='onstage'               #function to register items  
            else:                         #itemid is the primary key
                print()
                print('Invalid Choice')
                print('Making ITEM_TYPE As Null  ')
                d= 'Null'
            sh='show tables like "items"'
            c.execute(sh)
            fe=c.fetchone()       #checking if the table exists if not creating the table
            if fe is None:              
                cr='create table items(ITEM_ID numeric primary key,ITEM_NAME varchar(50) unique,ITEM_type varchar(25))'
                c.execute(cr)
            we='select ITEM_ID from items'       
            c.execute(we)
            fee=c.fetchall()
            val=False
            for i in fee:
                if i[0]==a:
                    print()     #preventing duplicated Entries
                    print('  Duplicate Item ID Detected')
                    print('  Registration Failed,Try Again!')
                    val=True
            if val==False:
                e="insert into items values({},'{}','{}')".format(a,b,d)
                c.execute(e)
                z.commit()
                print()
                print('  Successfully Registered')
            
            
                
        def update():
            g=int(input('Enter the ITEM ID to update:'))
            se='select * from items where ITEM_ID=({})'.format(g)
            c.execute(se)
            f=c.fetchone()      #function to update the registered items 
            val=False            
            if f is None:     
                print()      #checking if the itemid user provided exists!
                print('No Item Found Check ITEM_ID')
                val=True
            if val==False:
                  print()
                  print("    ITEM_ID   = ",f[0])
                  print('    ITEM_NAME = ',f[1])
                  print('    ITEM_type = ',f[2])
                  print()
                  h=input('DO you want to update( YES/NO)::')
                  if h[0] in "yY":
                     try:
                         a=int(input('Enter Item ID:'))
                         b=input('Enter Item Name:')
                         print('press 1 for offstage type:')
                         print('press 2 for onstage type:')
                         ch=int(input('Enter your choice ( 1 or 2):'))
                         if ch==1:
                           d='offstage'
                         elif ch==2:
                           d='onstage'         #updating the values
                         else:
                           print()
                           print('ITEM _TYPE selected is null')
                           d='NULL'
                         up="update items set ITEM_Id=({}),ITEM_NAME=('{}'),ITEM_type=('{}') where ITEM_ID=({})".format(a,b,d,g)
                         c.execute(up)
                         z.commit()
                         print()
                         print('updated successfully')
                     except mc.Error as e:
                       if e.errno == 1062:    #if an item is already registered with the Itemid which user Provides during updation  
                             print()        #preventing user to update the values
                             print('New ITEM_ID Provided Already Exists!!')
                             print('Check ITEM ID') 
                         
                  if h[0] not in "YyNn":
                     print()
                     print('Invalid Entry(Enter Yes/No)')
          


        def delete():
            g=int(input('Enter the ITEM ID to delete:'))
            se='select * from items where ITEM_ID=({})'.format(g)
            c.execute(se)                  #function to delete the registered items
            f=c.fetchone()
            if f is None:
                print()   #checking if the itemid user provided exists
                print('  Item Doesnt Exist(check ITEM_ID)')
            elif len(f)>0:
                  print()
                  print("    ITEM_ID   = ",f[0])
                  print('    ITEM_NAME = ',f[1])      
                  print('    ITEM_type = ',f[2])
                  print()
                  h=input('DO you want to delete( YES/NO)::')
                  if h[0] in "yY":
                    de='delete from items where ITEM_ID=({})'.format(g)
                    c.execute(de)   # deleting the item from the table
                    z.commit()
                    print()
                    print('Successfully deleted')
                  elif h[0] not in "YyNn":
                    print()
                    print(' Invalid Entry, Enter (yes/no)')


        def showall(): # function to show all the registered items in an excel file
            s = 'select * from items'
            c.execute(s)   # creating a new file items, fetching item information from SQL, and writing it into a CSV file
            fe = c.fetchall()
            if len(fe) < 1:
                print()
                print('There are no items to show!!!')
                return
            tp = "{}\\EXCEL TEMPLATES\\ITEMS(Template).xlsx".format(folder_path)
            workbook = load_workbook(tp)
            sheet = workbook.active
            sheet.cell(row = 2 , column = 2,value = schoolname)
            sheet.cell(row = 3 , column = 2,value = arname)
            logo = Image(logo_path)  #For adding school logo in the excel file
            logo.width = 118
            logo.height = 118           
            sheet.add_image(logo,'F2')
            template_row = sheet[7]
            cur_row = 7
            for row in fe:
                sheet.insert_rows(cur_row+1)
                for col_num in range(1, len(template_row) +1):
                    for attr in ['font', 'border', 'fill', 'number_format', 'protection', 'alignment']:
                        setattr(sheet.cell(row=cur_row+1 , column=col_num), attr, copy(getattr(template_row[col_num -1], attr)))
                sheet.cell(row=cur_row, column=3, value=row[0])
                sheet.cell(row=cur_row, column=4, value=row[1])
                sheet.cell(row=cur_row, column=5, value=row[2])
                cur_row += 1
            sheet.delete_rows(cur_row, 1)    
            mpath = os.path.join(excel_path,"ITEMS.xlsx")
            workbook.save(mpath)
            print()
            print('Opening The Document')
            os.startfile(mpath)
            # finding the path of the file
            # to start the excel automatically




        def sregister():
            a=int(input('Enter student ID:'))
            b=input('Enter student Name:')    #function to register students for different items 
            d=input('Enter class:')             
            k=int(input('Enter itemid of the item:'))   #studentid prevents duplicated entries 
            sh='show tables like "participants";'      #the table participants is joined with table items by Itemid 
            c.execute(sh)
            fe=c.fetchone()   
            if fe is None:
                cr='create table participants(STUDENT_ID numeric primary key,STUDENT_NAME varchar(50) ,CLASS varchar(10),ITEM_ID Numeric)'
                c.execute(cr)         #checking if table participants exists if not creating a new table
            m='select * from items'
            c.execute(m)
            ef=c.fetchall()
            q='select STUDENT_ID from participants'
            c.execute(q)
            eff=c.fetchall()
            val=False
            for j in eff:
                if j[0]==a:
                    print()         #preventing duplicated entries for studentid
                    print('Duplicate StudentID detected')
                    print('Registration Failed!Try again')
                    val=True
            if val==False:
                    valid=False
                    for i in ef:
                       if i[0]==k:    #inserting the values 
                          e="insert into participants values({},'{}','{}',{})".format(a,b,d,k)
                          c.execute(e)
                          z.commit()
                          print()
                          print('Successfully Registered')
                          valid=True

                    if valid==False:   
                        print()  #registration will fail if the itemid user provided doesnt exist
                        print('Registration failed Item doesnt exist (check ITEMID)')



        def supdate():
            g=int(input('Enter the STUDENT ID to update:'))
            n='select STUDENT_ID from participants'
            c.execute(n)             #function to update the registered students
            rf=c.fetchall()
            val=False
            for i in rf:
                if i[0]==g:
                    se='select * from participants where STUDENT_ID=({})'.format(g)
                    c.execute(se)
                    f=c.fetchone()   #displaying the student information 
                    print()
                    print("   STUDENT_ID   = ",f[0])
                    print('   STUDENT_NAME = ',f[1])
                    print('   CLASS        = ',f[2])
                    print('   ITEM_ID      = ',f[3])
                    print()
                    h=input('DO you want to update(YES/NO)::')
                    if h[0] in 'Nn':
                        val=True
                    elif h[0] not in "YyNn":
                        print()
                        print('Invalid entry enter (yes/no)')
                        val=True
                    elif h[0] in "yY":
                        try:
                            a=int(input('Enter student ID:'))
                            b=input('Enter student Name:')
                            d=input('Enter class:')
                            k=int(input('Enter itemid of the item:'))
                            m='select ITEM_ID from items'
                            c.execute(m)
                            ef=c.fetchall()
                            valid=False      #updating the values
                            for j in ef:
                                if j[0]==k:
                                    up="update participants set STUDENT_ID=({}),STUDENT_NAME=('{}'),CLASS=('{}'),ITEM_ID=({}) where STUDENT_ID=({})".format(a,b,d,k,g)
                                    c.execute(up)
                                    z.commit()
                                    print()
                                    print('Updated Successfully ')
                                    valid=True
                                      
                            if valid==False:
                                print()      #during updating if user provide itemid which doesnt exists raising an error
                                print('Updation failed Item doesnt exist (check ITEMID)')
                            val=True
                        except mc.Error as e:
                          if e.errno == 1062:   
                             print()   #checking if the any other students exists with studentid which user provided during updation
                             print(' New STUDENT_ID Provided Already Exists!!')
                             print(' Check STUDENT ID')
                             val=True
            if val==False:
                print()  #showing a message if the user provided studentid doesnt exist
                print(' Student doesnt exit (check studentid)')


        def sdelete():
            g=int(input('Enter the STUDENT ID to delete:'))
            n='select STUDENT_ID from participants'
            c.execute(n)     #function to delete the registered students
            rf=c.fetchall()
            val=False
            for i in rf:        
                if i[0]==g:
                    se='select * from participants where STUDENT_ID=({})'.format(g)
                    c.execute(se)
                    f=c.fetchone() #displaying student information
                    print()
                    print("   STUDENT_ID   = ",f[0])
                    print('   STUDENT_NAME = ',f[1])
                    print('   CLASS        = ',f[2])
                    print('   ITEM_ID      = ',f[3])
                    val=True
                    print()
                    ch=input(' Do You want to delete(yes/no):')                      
                    if ch[0] not in "YyNn":
                        print()
                        print('  Invalid entry enter (yes/no)')               
                    elif ch[0] in "yY":
                        h='delete from participants where STUDENT_ID=({})'.format(g)
                        c.execute(h)# deleting the participant from the table
                        z.commit()
                        print()
                        print('   Sucessfully deleted')
            if val==False:
                print()  #showing a message if the user provided studentid doesnt exist
                print('   Deletion Failed Student doesnt exist')
                print('   check STUDENT_ID')


        def showalls():  #function to show registered participants in an excel file
            s=' select p.STUDENT_ID,p.STUDENT_NAME,p.CLASS,i.ITEM_ID,i.ITEM_NAME,i.ITEM_type from participants as p,items as i where p.ITEM_ID=i.ITEM_ID;'
            c.execute(s)
            fe=c.fetchall()  #creating a file participants,fetching participant information from mysql writing into the file 
            if len(fe)<1:
                print()
                print('There are no participants to show!!!')
                return
            tp = "{}\\EXCEL TEMPLATES\\PARTICIPANTS(Template).xlsx".format(folder_path)
            workbook = load_workbook(tp)
            sheet = workbook.active
            sheet.cell(row = 2 , column = 2,value = schoolname)
            sheet.cell(row = 3 , column = 2,value = arname)
            logo = Image(logo_path)
            logo.width = 215  #For adding school logo in the excel file
            logo.height = 215
            sheet.add_image(logo,'G2')
            template_row = sheet[7]
            cur_row = 7  
            for row in fe:
                sheet.insert_rows(cur_row+1)
                for col_num in range(1, len(template_row) +1): #Inserting the Style of the Template
                    for attr in ['font', 'border', 'fill', 'number_format', 'protection', 'alignment']:
                        setattr(sheet.cell(row=cur_row+1 , column=col_num), attr, copy(getattr(template_row[col_num -1], attr)))
                sheet.cell(row=cur_row, column=2, value=row[0])
                sheet.cell(row=cur_row, column=3, value=row[1])
                sheet.cell(row=cur_row, column=4, value=row[2])
                sheet.cell(row=cur_row, column=5, value=row[3])#Inserting values to each columns
                sheet.cell(row=cur_row, column=6, value=row[4])
                sheet.cell(row=cur_row, column=7, value=row[5])
                cur_row+=1
            sheet.delete_rows(cur_row,1)
            mpath = os.path.join(excel_path,"PARTICIPANTS.xlsx")
            workbook.save(mpath)#Saving Excel File
            print()
            print('Opening The Document')
            os.startfile(mpath)  #to automatically open the excel file 


             
        def upload():
            sh='show tables like "winners"'
            c.execute(sh)   #function to upload winners of the items 
            fe=c.fetchone()
            if fe is None:   #checking if a table winners exists if not creating a new table winners
                cr='create table winners(ITEM_ID numeric unique,First varchar(50),Second varchar(50),Third varchar(50))'
                c.execute(cr)
            g=int(input('Enter the ITEM ID :'))
            se='select * from items where ITEM_ID=({})'.format(g)
            c.execute(se)
            f=c.fetchone()
            ry='select Item_id from winners'
            c.execute(ry)
            ryy=c.fetchall()
            var=False
            for j in ryy:
                if j[0]==g:   #checking if winners are uploaded or not for the itemid which user provides 
                    print()
                    print('Winners are already uploaded For This Item')
                    print('choose modify to update')
                    var=True
            if var==False:
                if f is None:
                    print()  #showing a message if item doesnt exist 
                    print('ITEM Doesnt exist check ITEM_ID')
                elif f is not None:
                    print()
                    print("   ITEM_ID   = ",f[0])
                    print('   ITEM_NAME = ',f[1])  #displaying the item 
                    print('   ITEM_TYPE = ',f[2])
                    see='select STUDENT_ID from participants where item_id=({})'.format(g)
                    c.execute(see)
                    sse=c.fetchall()
                    a=int(input('Enter the STUDENT ID who secured first:'))
                    b=int(input('Enter the STUDENT ID Who secured second:'))  #collecting studentid's who won 
                    l=int(input('Enter the STUDENT ID who secured Third:'))
                
                    pi=[int(i[0]) for i in sse]
                    if a not in pi or b not in pi or l not in pi:
                        print()  #checking if the studentid's provided are the participants of the itemid which user provides 
                        print('Student Entered is not a participant of the item')
                        print('Check STUDENT_ID !!')
                    elif a in pi and b in pi and l in pi:
                    
                        aa='select STUDENT_NAME from participants where STUDENT_ID=({})'.format(a)
                        bb='select STUDENT_NAME from participants where STUDENT_ID=({})'.format(b)
                        ll='select STUDENT_NAME from participants where STUDENT_ID=({})'.format(l)
                        c.execute(aa)
                        aaa=c.fetchone()
                        for i in aaa:      #fetching the names of students 
                          v=i
                        c.execute(bb)
                        bbb=c.fetchone()
                        for j in bbb:
                          vv=j
                        c.execute(ll)
                        lll=c.fetchone()
                        for k in lll:
                          vvv=k
                        d="insert into winners values({},'{}','{}','{}')".format(g,v,vv,vvv)
                        c.execute(d)
                        z.commit()   #inserting winners 
                        print()
                        print('Successfully uploaded')
                    


              
             
        def modify():
            g=int(input('Enter the ITEM ID to modify :'))
            se='select * from winners where ITEM_ID=({})'.format(g)
            c.execute(se)
            f=c.fetchone()   #function to modify uploaded winners 
            if f is None:
                print()
                print('Winners Are not uploaded')
            elif f is not None:
                    print()
                    print("  ITEM_ID  = ",f[0])
                    print('  FIRST    = ',f[1])
                    print('  SECOND   = ',f[2])   #printing winners info for the items 
                    print('  THIRD    = ',f[3])
                    x=input('DO you want to modify (yes/no)::')
                    if x[0] in 'Yy':
                       a=int(input('Enter the STUDENT ID who secured first:'))
                       b=int(input('Enter the STUDENT ID Who secured second:'))
                       l=int(input('Enter the STUDENT ID who secured Third:'))
                       see='select STUDENT_ID from participants where item_id=({})'.format(g)
                       c.execute(see)
                       sse=c.fetchall()
                       pi=[int(i[0]) for i in sse]
                       if a not in pi or b not in pi or l not in pi:
                         print() #checking if the studentid's provided are the participants of the itemid which user provides 
                         print('Student Entered is not a participant of the item')
                         print('Check STUDENT_ID !!')
                       elif a in pi and b in pi and l in pi:
                       
                         aa='select STUDENT_NAME from participants where STUDENT_ID=({})'.format(a)
                         bb='select STUDENT_NAME from participants where STUDENT_ID=({})'.format(b)
                         ll='select STUDENT_NAME from participants where STUDENT_ID=({})'.format(l)
                         c.execute(aa)
                         aaa=c.fetchone()
                         for i in aaa:
                           v=i
                         c.execute(bb)
                         bbb=c.fetchone()
                         for j in bbb:     #fetching studentname 
                           vv=j
                         c.execute(ll)
                         lll=c.fetchone()
                         for k in lll:
                           vvv=k
                         up="update winners set First=('{}'),Second=('{}'),Third=('{}') where ITEM_ID=({})".format(v,vv,vvv,g)
                         c.execute(up)
                         z.commit()
                         print()      #updating the values 
                         print('Modified successfully')

                    elif x[0] not in "yYnN":
                       print()
                       print('Invalid Entry Enter (yes/no)')
                

       
        def showwinners():  #function to show all the winners of items in an excel file 
            s=' select i.ITEM_NAME,i.ITEM_ID,w.First,w.Second,w.Third from winners as w,items as i where i.ITEM_ID=w.ITEM_ID'
            c.execute(s)   #creating a new file winners,fetching winners info from mysql and writing into the file 
            fe=c.fetchall()
            if len(fe)<1:
                print()
                print('There are no Winners to show!!!')
                return
            tp = "{}\\EXCEL TEMPLATES\\WINNERS(Template).xlsx".format(folder_path)
            workbook = load_workbook(tp)
            sheet = workbook.active
            sheet.cell(row = 2 , column = 2,value = schoolname)
            sheet.cell(row = 3 , column = 2,value = arname)
            logo = Image(logo_path)
            logo.width = 252
            logo.height = 252
            sheet.add_image(logo, 'F2')
            template_row = sheet[7]
            cur_row = 7
            for row in fe:
                sheet.insert_rows(cur_row+1)
                for col_num in range(1, len(template_row) +1):
                    for attr in ['font', 'border', 'fill', 'number_format', 'protection', 'alignment']:
                        setattr(sheet.cell(row=cur_row+1 , column=col_num), attr, copy(getattr(template_row[col_num -1], attr)))
                sheet.cell(row=cur_row, column=2, value=row[0])
                sheet.cell(row=cur_row, column=3, value=row[1])
                sheet.cell(row=cur_row, column=4, value=row[2])#For adding school logo in the excel file
                sheet.cell(row=cur_row, column=5, value=row[3])
                sheet.cell(row=cur_row, column=6, value=row[4])
                cur_row += 1
            sheet.delete_rows(cur_row, 1)  
            mpath = os.path.join(excel_path,"WINNERS.xlsx")
            workbook.save(mpath)
            print()
            print('Opening The Document')
            os.startfile(mpath) #to automatically open the excel file         
            



        def certificate():
            logo = cv2.imread(logo_path)  #resizing the school logo
            logo = cv2.resize(logo, (150, 150))
            print()
            print('╔══════════════════════════════╗')
            print("*                              *")
            print("* 1-Participation Certificate  *")   # asking whether user needs participant's /winner's certificate
            print("* 2-Winner's certificate       *")
            print("*                              *")
            print('╚══════════════════════════════╝')
            n = int(input("Enter your choice (1 or 2): "))
            if n > 2:
                print()
                print("Invalid choice choose (1 or 2)")
            elif n == 2:
                a = int(input("Enter the ITEM_ID: "))
                s = 'select * from items where ITEM_ID=({})'.format(a)
                c.execute(s)
                f = c.fetchone()
                if f is None:    #checking if the itemid user provided exists
                    print("Item Doesn't exist")
                elif f is not None:
                      
                  print()
                  print("    ITEM_ID   = ",f[0])
                  print('    ITEM_NAME = ',f[1])
                  print('    ITEM_type = ',f[2])
                  print()
                  h=int(input('Press 1 To Open the Certificates::'))
                  if h>1:
                      print()
                      print('Invalid Choice!,Cancelling Generation')
                  if h==1:
                    ns=f[1]
                    nsp=os.path.join(wcert_path,ns)
                    if not os.path.exists(nsp):    #fetching the information of winners from mysql
                        os.makedirs(nsp)
                    se='select w.first,p.class,i.item_name from items as i, winners as w,participants as p where i.item_id=w.ITEM_ID=p.item_id and w.Item_id=({})'.format(a)
                    sel='select w.second,p.class,i.item_name from items as i, winners as w,participants as p where i.item_id=w.ITEM_ID=p.item_id and w.Item_id=({})'.format(a)
                    sele='select w.third,p.class,i.item_name from items as i, winners as w,participants as p where i.item_id=w.ITEM_ID=p.item_id and w.Item_id=({})'.format(a)
                    c.execute(se)
                    fe = c.fetchall()
                    for i in fe:
                        wcert = cv2.imread('{}\\template\\FIRST.jpg'.format(folder_path)) #reading the template of the certificate
                        if logo is not None:
                            wcert[10:10 + logo.shape[1], 618:618 + logo.shape[1]]=logo
                        cv2.putText(wcert,i[0], (544,803), cv2.FONT_HERSHEY_TRIPLEX, 1.5, (0, 0, 255), 2, cv2.LINE_AA)
                        cv2.putText(wcert,i[1], (1290,803), cv2.FONT_HERSHEY_TRIPLEX, 1.5, (0, 0, 255), 2, cv2.LINE_AA)                    
                        cv2.putText(wcert,i[2], (1002,871), cv2.FONT_HERSHEY_TRIPLEX, 1.5, (0, 0, 255), 2, cv2.LINE_AA)                    
                        cv2.putText(wcert,schoolname, (770,122), cv2.FONT_HERSHEY_COMPLEX, 2.7, (255, 0, 0), 3, cv2.LINE_AA)
                        cv2.putText(wcert,arname, (503,364), cv2.FONT_HERSHEY_COMPLEX, 2.2, (0, 0, 0), 3, cv2.LINE_AA)
                        cv2.imwrite(r"{}\\cert\\winners_certificates\\{}\\{}({}).jpg".format(folder_path,ns,i[0],i[2]),wcert)
                    c.execute(sel)   # writing the values into the template 
                    fet=c.fetchall()
                    for i in fet:
                        wcert = cv2.imread('{}\\template\\SECOND.jpg'.format(folder_path))
                        if logo is not None:
                                wcert[10:10 + logo.shape[1], 618:618 + logo.shape[1]]=logo
                        cv2.putText(wcert,i[0], (544,803), cv2.FONT_HERSHEY_TRIPLEX, 1.5, (0, 0, 255), 2, cv2.LINE_AA)
                        cv2.putText(wcert,i[1], (1290,803), cv2.FONT_HERSHEY_TRIPLEX, 1.5, (0, 0, 255), 2, cv2.LINE_AA)                    
                        cv2.putText(wcert,i[2], (1002,871), cv2.FONT_HERSHEY_TRIPLEX, 1.5, (0, 0, 255), 2, cv2.LINE_AA)                    
                        cv2.putText(wcert,schoolname, (770,122), cv2.FONT_HERSHEY_COMPLEX, 2.7, (255, 0, 0), 3, cv2.LINE_AA)
                        cv2.putText(wcert,arname, (503,364), cv2.FONT_HERSHEY_COMPLEX, 2.2, (0, 0, 0), 3, cv2.LINE_AA)
                        cv2.imwrite(r"{}\\cert\\winners_certificates\\{}\\{}({}).jpg".format(folder_path,ns,i[0],i[2]),wcert)
                    c.execute(sele)
                    fetc=c.fetchall()
                    for i in fetc:
                        wcert = cv2.imread('{}\\template\\THIRD.jpg'.format(folder_path))
                        if logo is not None:
                                wcert[10:10 + logo.shape[1], 618:618 + logo.shape[1]]=logo
                        cv2.putText(wcert,i[0], (544,803), cv2.FONT_HERSHEY_TRIPLEX, 1.5, (0, 0, 255), 2, cv2.LINE_AA)
                        cv2.putText(wcert,i[1], (1290,803), cv2.FONT_HERSHEY_TRIPLEX, 1.5, (0, 0, 255), 2, cv2.LINE_AA)                    
                        cv2.putText(wcert,i[2], (1002,871), cv2.FONT_HERSHEY_TRIPLEX, 1.5, (0, 0, 255), 2, cv2.LINE_AA)                    
                        cv2.putText(wcert,schoolname, (770,122), cv2.FONT_HERSHEY_COMPLEX, 2.7, (255, 0, 0), 3, cv2.LINE_AA)
                        cv2.putText(wcert,arname, (503,364), cv2.FONT_HERSHEY_COMPLEX, 2.2, (0, 0, 0), 3, cv2.LINE_AA)
                        cv2.imwrite(r"{}\\cert\\winners_certificates\\{}\\{}({}).jpg".format(folder_path,ns,i[0],i[2]),wcert)
                    if os.path.exists("..\\cert"):
                        print()
                        print('Opening The Folder')
                        cp = os.path.abspath("..\\cert")   
                        sb.Popen(['explorer', cp]) #to automatically open the folder 
                    
                                
            elif n==1:
                ab = int(input("Enter the ITEM_ID: "))
                qp = 'select * from items where ITEM_ID=({})'.format(ab)
                c.execute(qp)
                ti = c.fetchone()
                if ti is None:
                    print("Item Doesn't exist check ITEM_ID")
                elif ti is not None:
                      print()
                      print("    ITEM_ID   = ",ti[0])
                      print('    ITEM_NAME = ',ti[1])
                      print('    ITEM_type = ',ti[2])
                      print()
                      h=int(input('Press 1 To Open the Certificates::'))
                      if h>1:
                          print()
                          print('Invalid Choice!,Cancelling Generation')
                      if h==1:
                         sn=ti[1]
                         snp=os.path.join(pcert_path,sn)
                         if not os.path.exists(snp):
                            os.makedirs(snp)
                         que = 'select p.STUDENT_name,p.class,i.item_name from items as i,participants as p where i.item_id=p.item_id and p.Item_id=({})'.format(ab)
                         c.execute(que)
                         fet = c.fetchall()
                         for j in fet:
                                pcert = cv2.imread('{}\\template\\participant.jpg'.format(folder_path))
                                if logo is not None:
                                    pcert[10:10 + logo.shape[1], 620:620 + logo.shape[1]] = logo
                                cv2.putText(pcert,j[0], (544,803), cv2.FONT_HERSHEY_TRIPLEX, 1.5, (0, 0, 255), 2, cv2.LINE_AA)
                                cv2.putText(pcert,j[1], (1290,803), cv2.FONT_HERSHEY_TRIPLEX, 1.5, (0, 0, 255), 2, cv2.LINE_AA)                    
                                cv2.putText(pcert,j[2], (1008,870), cv2.FONT_HERSHEY_TRIPLEX, 1.5, (0, 0, 255), 2, cv2.LINE_AA)                    
                                cv2.putText(pcert,schoolname,(770,122), cv2.FONT_HERSHEY_COMPLEX, 2.7, (255, 0, 0), 3, cv2.LINE_AA)
                                cv2.putText(pcert,arname, (503,364), cv2.FONT_HERSHEY_COMPLEX, 2.2, (0, 0, 0), 3, cv2.LINE_AA)
                                cv2.imwrite(r"{}\\cert\\participation_certificates\\{}\\{}({}).jpg".format(folder_path,sn,j[0],j[2]),pcert)
                         if os.path.exists("..\\cert"):     
                            print()
                            print('Opening The Folder')
                            cp = os.path.abspath("..\\cert")
                            sb.Popen(['explorer', cp])
                    
                        

                                                                
        print()
        print('                                  ' )
        
        print('                          ',schoolname)  #school & arts fest name collected in the begining of program will come here !
        print('                          ',arname)
        print('                 ********* Welcome To Arts planner!!!*********')
        while True:
            print()
            print('╔════════════════**********═══════════════╗')
            print('║              PROGRAM OPTIONS            ║')
            print('║             _________________           ║')
            print('║                                         ║')
            print('║        ITEMS                            ║')
            print('║        1 ->  Register An Item           ║')
            print('║        2 ->  Update An Item             ║')
            print('║        3 ->  Delete An Item             ║')
            print('*        4 ->  Show All Items             *')
            print('*                                         *')
            print('*        PARTICIPANTS                     *')
            print('*        5 ->  Register A Participant     *')
            print('*        6 ->  Update A Participant       *')
            print('║        7 ->  Delete A Participant       ║')
            print('║        8 ->  Show All Participants      ║')
            print('║                                         ║')
            print('║        WINNERS                          ║')
            print('║        9 ->  Upload Winners             ║')
            print('║       10 ->  Modify Winners             ║')
            print('║       11 ->  Show All Winners           ║')
            print('║       12 ->  Generate Certificates      ║')
            print('║                                         ║')
            print('║       13 ->  Exit                       ║')
            print('║                                         ║')
            print('║             *****************           ║')
            print('╚════════════════**********═══════════════╝')
            print()
            print('           Press The Corresponding Keys ')
            print()

            cho=int(input('     Enter Your Choice::'))
            if cho>13:
                print()
                print('Invalid choice')
            elif cho==1:
                register()       #running the functions as per user's choice 
            elif cho==2:
                update()
            elif cho==3:
                delete()
            elif cho==4:
                showall()
            elif cho==5:
                sregister()
            elif cho==6:
                supdate()
            elif cho==7:
                sdelete()
            elif cho==8:
                showalls()
            elif cho==9:
                upload()
            elif cho==10:
                modify()
            elif cho==11:
                showwinners()
            elif cho==12:
                certificate()
            elif cho==13:
                print()
                print('╔═════════════════════════════════════════╗')
                print("║             ~~~THANK YOU~~~             ║")
                print("╚═════════════════════════════════════════╝")
                break
        break
        z.close() # closing the established connection

    except ImportError as ie:
        print() #showing an error message if the required modules are not imported
        print("+======================================================+")
        print("!   *Ensure That You Have Installed required modules   !")
        print("               ",f' *{ie}',"                            ")                #showing which module is not installed  
        print("!                    *Try Again!!!                     !")
        print("+======================================================+")
        break
    except mc.Error as e:
        if e.errno == 1045:
            print()   #checking authentication preventing the user to run the program if user fails to give the password
            print("+=====================+")
            print("!     LOGIN FAILED    !")
            print("!    WRONG PASSWORD!  !")
            print("!      TRY AGAIN!!    !")
            print("+=====================+")
            continue
        if e.errno == 1064:
            print()   #showing a message if user provides special/invalid characters as names for database
            print("+=============================================================+")
            print("!                Failed To Create A Database!!!               !")
            print("!           *Database Name Must begin With a letter           !")
            print("!        *Dont Use Special Character's Except Underscore      !")
            print("!     *Ensure that The Name Is Not More Than 64 Characters    !")
            print("!               *****Follow The Above Rules*****              !")
            print("!                        Try Again!!!!                        !")
            print("+=============================================================+")
            continue
        if e.errno == 1146:
            print() #showing a message if user  clicks show all item/participants/winners without adding thier info
            print("+===============================+")
            print("!        Table Not Found        !")  #showing a message that the tabel doesnt exist
            print("!   Perform Registration First  !")
            print("!          Try Again!!!!        !")
            print("+===============================+")
            continue
    except ValueError:
        print() #if user provided invalid/special characters when entering choices 
        print('Its An Invalid Choice')
        continue
    except PermissionError:
        print()#while keeping the excel file open if user tries to open the file by clicking choices show a message for permission error
        print("+========================================+")
        print("!            Permission Denied           !")
        print("!   *Ensure that You have Added Records  !") 
        print("!   *Close the File If You Are Using It  !")
        print("!      ****Check The Above Rules****     !")
        print("!               *Try Again*              !")
        print("+========================================+")
        continue
    except FileNotFoundError:
        print() #showing an error message if the file not Found
        print("+==================================================+")
        print("!                  File Not Found                  !")
        print("!   *Ensure That You Have Performed Registrations  !")
        print("!          *Dont Delete The Files Manually         !")
        print("!                   *Try Again!!!!                 !")
        print("+==================================================+")
        continue
    except NameError:
        print()
        print(' Try Again!!!!')
        continue
    except NotADirectoryError:
        root = tk.Tk()
        root.attributes('-topmost',True)        
        root.withdraw()                          #showing the messagebox to give error message
        a = messagebox.showerror('SCHOOL ARTS PLANNER','Directory Not Found!!\nTry Again!!!')
        root.destroy()
        continue                
       

    
        
            
    
    
   
    

    
   
    

        
    
    

    
    
    

